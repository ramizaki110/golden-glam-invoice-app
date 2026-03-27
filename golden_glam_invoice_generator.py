import os
import base64
import tempfile
from pathlib import Path
from datetime import datetime
from collections import deque

from PIL import Image as PILImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
LOGO_PATH = BASE_DIR / "golden_glam_logo_final.png"

styles = getSampleStyleSheet()

BLACK = colors.HexColor("#231f1e")
DARK = colors.HexColor("#4a4745")
MID = colors.HexColor("#8c8a87")
BORDER = colors.HexColor("#d8d5d2")

CONTENT_WIDTH = 7.30 * inch
ITEM_COL_WIDTHS = [
    0.66 * inch,  # item no
    2.10 * inch,  # description
    0.92 * inch,  # est del
    0.52 * inch,  # type
    0.34 * inch,  # qty
    0.72 * inch,  # unit price
    0.42 * inch,  # disc
    0.58 * inch,  # total
    1.04 * inch,  # photo
]
assert round(sum(ITEM_COL_WIDTHS), 4) == round(CONTENT_WIDTH, 4)


def usd(v: float) -> str:
    return f"${v:,.0f}"


def fmt_date_for_footer(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return date_str or ""


def _is_dark_pixel(px, threshold=60):
    r, g, b, a = px
    return a > 0 and r <= threshold and g <= threshold and b <= threshold


def _whiten_edge_connected_dark(img_rgba: PILImage.Image, threshold=60) -> PILImage.Image:
    """
    Flood-fill from the image edges and turn dark edge-connected pixels white.
    This fixes common product cutouts that arrive with black canvas background.
    """
    img = img_rgba.copy()
    w, h = img.size
    px = img.load()

    visited = set()
    q = deque()

    for x in range(w):
        q.append((x, 0))
        q.append((x, h - 1))
    for y in range(h):
        q.append((0, y))
        q.append((w - 1, y))

    while q:
        x, y = q.popleft()
        if (x, y) in visited:
            continue
        visited.add((x, y))

        if x < 0 or y < 0 or x >= w or y >= h:
            continue

        if _is_dark_pixel(px[x, y], threshold):
            px[x, y] = (255, 255, 255, 255)
            for nx, ny in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                if 0 <= nx < w and 0 <= ny < h and (nx, ny) not in visited:
                    q.append((nx, ny))

    return img


def _decode_image(image_value: str) -> str | None:
    """
    Decode image and force a white background robustly:
    - preserve transparency if present
    - whiten dark edge-connected backgrounds
    - flatten onto white
    """
    if not image_value:
        return None

    try:
        if image_value.startswith("data:image"):
            _, b64 = image_value.split(",", 1)
            raw = base64.b64decode(b64)
            tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".img")
            tmp_in.write(raw)
            tmp_in.close()
            src_path = tmp_in.name
        else:
            if not os.path.exists(image_value):
                return None
            src_path = image_value

        img = PILImage.open(src_path).convert("RGBA")
        img = _whiten_edge_connected_dark(img, threshold=60)

        white_bg = PILImage.new("RGBA", img.size, (255, 255, 255, 255))
        white_bg.paste(img, (0, 0), img)

        out = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        white_bg.save(out.name, format="PNG")
        out.close()

        if image_value.startswith("data:image"):
            try:
                os.remove(src_path)
            except Exception:
                pass

        return out.name

    except Exception:
        return None


class _NumberedCanvas(canvas.Canvas):
    """Two-pass canvas that knows the total page count."""
    _gg_inv = {}  # set by draw_invoice before doc.build()

    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer_and_header(total)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def _draw_footer_and_header(self, total_pages):
        inv = _NumberedCanvas._gg_inv
        print_date = fmt_date_for_footer(inv.get("date", ""))
        page_num = self.getPageNumber()

        self.saveState()

        # ── Header: logo on every page, same size and position ──────────────
        if os.path.exists(str(LOGO_PATH)):
            logo_w, logo_h = 2.85 * inch, 1.08 * inch
            self.drawImage(
                str(LOGO_PATH),
                (letter[0] - logo_w) / 2,
                letter[1] - 0.3 * inch - logo_h,
                width=logo_w, height=logo_h,
                mask="auto", preserveAspectRatio=True,
            )

        # ── Footer ─────────────────────────────────────────────────────────────
        y_line = 0.70 * inch
        self.setStrokeColor(BORDER)
        self.setLineWidth(0.6)
        self.line(0.28 * inch, y_line, letter[0] - 0.28 * inch, y_line)

        self.setFont("Helvetica", 7)
        self.setFillColor(MID)
        self.drawString(0.28 * inch, y_line - 0.16 * inch, f"Print date: {print_date}")
        self.drawRightString(
            letter[0] - 0.28 * inch,
            y_line - 0.16 * inch,
            f"Page {page_num} of {total_pages}",
        )

        self.setFont("Helvetica-Bold", 7)
        self.setFillColor(BLACK)
        self.drawCentredString(letter[0] / 2, 0.46 * inch, "GOLDEN GLAM INTERIORS LLC")

        self.setFont("Helvetica", 6)
        self.setFillColor(DARK)
        self.drawCentredString(
            letter[0] / 2,
            0.32 * inch,
            "Address: 828 Highland Ln Ne, Apt. 2204, Atlanta, GA 30306  |  Phone: 770-375-7343",
        )
        self.drawCentredString(
            letter[0] / 2,
            0.20 * inch,
            "Bank #: 930283558  |  Routing: 061092387  |  Zelle: rana_salah@goldenglam.nl  |  E-mail: sales@goldenglam.nl",
        )
        self.restoreState()


def _footer(canvas, doc):
    """No-op: footer/header is handled by _NumberedCanvas."""
    pass


def _autosize(ws, widths=None):
    if widths:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        return

    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)


def _write_internal_excel(inv: dict, output_path: str):
    wb = Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    thin       = Side(style="thin", color="D9D9D9")
    thick_side = Side(style="medium", color="231F1E")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="231F1E")   # dark header
    gold_fill  = PatternFill("solid", fgColor="B8963E")   # gold accent
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    amber_fill = PatternFill("solid", fgColor="FFF3CD")
    red_fill   = PatternFill("solid", fgColor="F8D7DA")
    grey_fill  = PatternFill("solid", fgColor="F5F2EF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_font   = Font(color="FFFFFF", bold=True, size=9)
    bold_font  = Font(bold=True, size=9)
    norm_font  = Font(size=9)
    title_font = Font(bold=True, size=12, color="231F1E")
    gold_font  = Font(bold=True, size=10, color="FFFFFF")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right      = Alignment(horizontal="right",  vertical="center")

    def hdr_cell(ws, row, col, val):
        c = ws.cell(row, col, val)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = bdr
        return c

    def data_cell(ws, row, col, val, fmt=None, bold=False, fill=None, align=None):
        c = ws.cell(row, col, val)
        c.font = Font(bold=bold, size=9)
        c.alignment = align or left
        c.border = bdr
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c

    # ── Pre-compute financials ─────────────────────────────────────────────────
    items = inv.get("items", [])
    delivery_charge = inv.get("delivery_charge", 0) or 0
    tax_rate        = inv.get("tax_rate", 0) or 0

    item_data = []
    subtotal = 0
    total_cost = 0
    for item in items:
        qty        = item.get("qty", 0)
        unit_price = item.get("unit_price", 0)
        disc       = item.get("discount", 0)
        lt         = qty * unit_price * (1 - disc)
        unit_cost  = item.get("cost", 0) or 0
        ext_cost   = unit_cost * qty
        profit     = lt - ext_cost if unit_cost else None
        gm         = (profit / lt) if (profit is not None and lt) else None
        subtotal  += lt
        if unit_cost: total_cost += ext_cost
        item_data.append(dict(
            no=item.get("no",""), vendor_name=item.get("vendor_name",""),
            vendor_no=item.get("vendor_no",""), desc=item.get("description",""),
            qty=qty, unit=item.get("unit",""), unit_price=unit_price,
            disc=disc, lt=lt, raw_cost=item.get("raw_cost",0) or 0,
            cost_disc=item.get("cost_disc",0) or 0, unit_cost=unit_cost,
            ext_cost=ext_cost, profit=profit, gm=gm,
            delivery=item.get("delivery",""),
            has_img=bool(item.get("image")),
        ))

    tax_amt     = (subtotal + delivery_charge) * tax_rate
    grand_total = subtotal + delivery_charge + tax_amt
    total_profit = subtotal - total_cost if total_cost else None
    overall_gm   = (total_profit / subtotal) if (total_profit is not None and subtotal) else None

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — P&L Summary  (at-a-glance)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "P&L Summary"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 16
    ws1.row_dimensions[1].height = 28

    # Title banner
    ws1.merge_cells("A1:E1")
    tc = ws1.cell(1, 1, f"GOLDEN GLAM — P&L SUMMARY  |  Invoice {inv.get('number','')}  |  {inv.get('client_name','')}  |  {inv.get('date','')}")
    tc.font = Font(bold=True, size=11, color="FFFFFF")
    tc.fill = hdr_fill; tc.alignment = center

    # ── Section A: Overall P&L ────────────────────────────────────────────────
    r = 3
    ws1.merge_cells(f"A{r}:E{r}")
    sc = ws1.cell(r, 1, "OVERALL INVOICE P&L")
    sc.font = gold_font; sc.fill = gold_fill; sc.alignment = center
    ws1.row_dimensions[r].height = 20

    r += 1
    for col, h in enumerate(["", "Amount ($)", "% of Revenue", "", ""], 1):
        hdr_cell(ws1, r, col, h)
    ws1.row_dimensions[r].height = 18

    def pl_row(ws, r, label, value, pct=None, highlight=None):
        ws.row_dimensions[r].height = 18
        c1 = ws.cell(r, 1, label); c1.font = bold_font; c1.border = bdr; c1.fill = grey_fill; c1.alignment = left
        c2 = ws.cell(r, 2, value); c2.font = bold_font; c2.border = bdr; c2.alignment = right
        c2.number_format = "$#,##0.00"
        c3 = ws.cell(r, 3, pct);   c3.font = norm_font; c3.border = bdr; c3.alignment = right
        c3.number_format = "0.00%"   # always set — blank cells show nothing, decimal cells show %
        ws.cell(r, 4).border = bdr; ws.cell(r, 5).border = bdr
        if highlight: c2.fill = highlight; c3.fill = highlight

    r += 1; pl_row(ws1, r, "Invoice Subtotal (Revenue)", subtotal)
    r += 1; pl_row(ws1, r, "Total Cost of Goods", total_cost if total_cost else "N/A",
                   (total_cost/subtotal) if (total_cost and subtotal) else None)
    r += 1
    if total_profit is not None:
        fill = green_fill if total_profit >= 0 else red_fill
        pl_row(ws1, r, "Gross Profit (items)", total_profit,
               overall_gm, highlight=fill)
    else:
        pl_row(ws1, r, "Gross Profit (items)", "N/A — costs not entered")
    r += 1; pl_row(ws1, r, "Delivery Charged to Client", delivery_charge)
    r += 1; pl_row(ws1, r, "Sales Tax", tax_amt, tax_rate)
    r += 1
    ws1.merge_cells(f"A{r}:E{r}")
    note = ws1.cell(r, 1, "ℹ  Delivery P&L (net profit on delivery) → see 'Delivery P&L' tab")
    note.font = Font(italic=True, size=8, color="888888"); note.alignment = left

    # ── Section B: Line-item P&L ──────────────────────────────────────────────
    r += 2
    ws1.merge_cells(f"A{r}:E{r}")
    sc2 = ws1.cell(r, 1, "LINE ITEM BREAKDOWN")
    sc2.font = gold_font; sc2.fill = gold_fill; sc2.alignment = center
    ws1.row_dimensions[r].height = 20

    r += 1
    for col, h in enumerate(["Item / Description", "Line Total ($)", "Cost ($)", "Profit ($)", "GM %"], 1):
        hdr_cell(ws1, r, col, h)
    ws1.row_dimensions[r].height = 18

    for d in item_data:
        r += 1
        ws1.row_dimensions[r].height = 16
        label = f"[{d['no']}] {d['desc'][:40]}" if d['no'] else d['desc'][:45]
        data_cell(ws1, r, 1, label, bold=False)
        data_cell(ws1, r, 2, d['lt'],       fmt="$#,##0.00", align=right)
        data_cell(ws1, r, 3, d['ext_cost'] if d['unit_cost'] else "—", fmt="$#,##0.00" if d['unit_cost'] else None, align=right)
        if d['profit'] is not None:
            pf = d['profit']
            fill = green_fill if pf >= 0 else red_fill
            data_cell(ws1, r, 4, pf,     fmt="$#,##0.00", align=right, fill=fill, bold=True)
            data_cell(ws1, r, 5, d['gm'] if d['gm'] is not None else "—",
                      fmt="0.00%", align=right,
                      fill=green_fill if d['gm'] and d['gm']>=0.30 else (amber_fill if d['gm'] and d['gm']>=0.15 else red_fill))
        else:
            data_cell(ws1, r, 4, "—", align=right)
            data_cell(ws1, r, 5, "—", align=right)

    # Totals row
    r += 1
    ws1.row_dimensions[r].height = 18
    for col in range(1, 6):
        ws1.cell(r, col).border = Border(top=thick_side, bottom=thick_side, left=thin, right=thin)
    data_cell(ws1, r, 1, "TOTAL", bold=True, fill=grey_fill)
    data_cell(ws1, r, 2, subtotal,      fmt="$#,##0.00", bold=True, align=right)
    data_cell(ws1, r, 3, total_cost if total_cost else "—", fmt="$#,##0.00" if total_cost else None, bold=True, align=right)
    if total_profit is not None:
        fill = green_fill if total_profit >= 0 else red_fill
        data_cell(ws1, r, 4, total_profit, fmt="$#,##0.00", bold=True, align=right, fill=fill)
        data_cell(ws1, r, 5, overall_gm,   fmt="0.00%",     bold=True, align=right,
                  fill=green_fill if overall_gm and overall_gm>=0.30 else (amber_fill if overall_gm and overall_gm>=0.15 else red_fill))
    else:
        data_cell(ws1, r, 4, "—", bold=True, align=right)
        data_cell(ws1, r, 5, "—", bold=True, align=right)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — Invoice Detail  (full line item breakdown)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Invoice Detail")
    det_headers = [
        "Item No", "Vendor Name", "Vendor No", "Description",
        "Qty", "Unit", "Unit Price", "Disc %",
        "Line Total", "Raw Cost", "Cost Disc %", "Unit Cost",
        "Ext. Cost", "Profit", "GM %", "Est. Delivery", "Photo",
    ]
    for col, h in enumerate(det_headers, 1):
        hdr_cell(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 32

    for i, d in enumerate(item_data):
        row_idx = i + 2
        ws2.row_dimensions[row_idx].height = 16
        vals = [
            d["no"], d["vendor_name"], d["vendor_no"], d["desc"],
            d["qty"], d["unit"], d["unit_price"], d["disc"],
            d["lt"], d["raw_cost"], d["cost_disc"]/100 if d["cost_disc"] else 0,
            d["unit_cost"], d["ext_cost"],
            d["profit"] if d["profit"] is not None else "N/A",
            d["gm"] if d["gm"] is not None else "N/A",
            d["delivery"], "Yes" if d["has_img"] else "",
        ]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row_idx, col, v)
            c.font = norm_font; c.border = bdr; c.alignment = left
        # Formats — col map: 5=Qty(plain), 7=UnitPrice($), 8=Disc%(%), 
        #   9=LineTotal($), 10=RawCost($), 11=CostDisc%(%), 12=UnitCost($),
        #   13=ExtCost($), 14=Profit($), 15=GM%(%)
        # Qty col: store as int, format as plain number
        qty_cell = ws2.cell(row_idx, 5)
        qty_cell.value = int(d["qty"])
        qty_cell.number_format = "0"
        for col, fmt in [(7,"$#,##0.00"),(9,"$#,##0.00"),(10,"$#,##0.00"),
                         (12,"$#,##0.00"),(13,"$#,##0.00"),(14,"$#,##0.00")]:
            ws2.cell(row_idx, col).number_format = fmt
        for col in [8, 11, 15]:   # all % columns
            c = ws2.cell(row_idx, col)
            c.number_format = "0.00%"
            # ensure value is stored as decimal for percentage display
            if isinstance(c.value, (int, float)):
                c.value = float(c.value)  # force float so Excel treats as number
        # GM colour
        gm_val = d["gm"]
        if gm_val is not None:
            ws2.cell(row_idx, 15).fill = (green_fill if gm_val>=0.30 else amber_fill if gm_val>=0.15 else red_fill)

    _autosize(ws2, {1:12, 2:18, 3:14, 4:36, 5:6, 6:8, 7:12, 8:9,
                    9:12, 10:12, 11:11, 12:12, 13:12, 14:12, 15:9, 16:20, 17:8})

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — Delivery P&L
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Delivery P&L")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 16
    ws3.row_dimensions[1].height = 28

    # Title
    ws3.merge_cells("A1:C1")
    t3 = ws3.cell(1, 1, f"DELIVERY P&L  |  Invoice {inv.get('number','')}  |  {inv.get('client_name','')}")
    t3.font = Font(bold=True, size=10, color="FFFFFF")
    t3.fill = hdr_fill; t3.alignment = center

    # Charged to client
    r = 3
    ws3.merge_cells(f"A{r}:C{r}")
    ws3.cell(r, 1, "DELIVERY CHARGED TO CLIENT").font = Font(bold=True, size=9)
    ws3.cell(r, 1).fill = grey_fill; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).alignment = left
    r += 1
    ws3.cell(r, 1, inv.get("delivery_type","Delivery") or "Delivery").border = bdr
    ws3.cell(r, 1).font = norm_font; ws3.cell(r, 1).alignment = left
    ws3.cell(r, 2, "").border = bdr
    c_charged = ws3.cell(r, 3, delivery_charge)
    c_charged.number_format = "$#,##0.00"; c_charged.border = bdr
    c_charged.font = Font(bold=True, size=9); c_charged.alignment = right

    # Delivery costs (manual entry)
    r += 2
    ws3.merge_cells(f"A{r}:C{r}")
    ws3.cell(r, 1, "YOUR DELIVERY COSTS  (fill in manually)").font = Font(bold=True, size=9)
    ws3.cell(r, 1).fill = grey_fill; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).alignment = left
    r += 1
    for col, h in enumerate(["From", "To", "Cost ($)"], 1):
        hdr_cell(ws3, r, col, h)
    ws3.row_dimensions[r].height = 18
    cost_start_row = r + 1
    for i in range(10):
        rr = r + 1 + i
        for col in range(1, 4):
            c = ws3.cell(rr, col, "")
            c.border = bdr; c.font = norm_font; c.alignment = left
            if col == 3:
                c.number_format = "$#,##0.00"; c.alignment = right
        ws3.row_dimensions[rr].height = 16
    cost_end_row = r + 10

    # Total delivery costs formula
    r = cost_end_row + 1
    ws3.row_dimensions[r].height = 18
    ws3.cell(r, 1, "Total Delivery Costs").font = bold_font; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).fill = grey_fill; ws3.cell(r, 1).alignment = left
    ws3.cell(r, 2, "").border = bdr
    total_cost_cell = ws3.cell(r, 3)
    total_cost_cell.value = f"=SUM(C{cost_start_row}:C{cost_end_row})"
    total_cost_cell.number_format = "$#,##0.00"; total_cost_cell.border = bdr
    total_cost_cell.font = bold_font; total_cost_cell.alignment = right

    # Delivery P&L result
    r += 2
    ws3.merge_cells(f"A{r}:C{r}")
    ws3.cell(r, 1, "DELIVERY P&L").font = gold_font
    ws3.cell(r, 1).fill = gold_fill; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).alignment = center
    ws3.row_dimensions[r].height = 20

    r += 1
    ws3.row_dimensions[r].height = 18
    ws3.cell(r, 1, "Charged to Client").font = norm_font; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).alignment = left
    ws3.cell(r, 2, "").border = bdr
    ws3.cell(r, 3, delivery_charge).number_format = "$#,##0.00"; ws3.cell(r, 3).border = bdr; ws3.cell(r, 3).alignment = right

    r += 1
    ws3.row_dimensions[r].height = 18
    ws3.cell(r, 1, "Total Delivery Costs").font = norm_font; ws3.cell(r, 1).border = bdr; ws3.cell(r, 1).alignment = left
    ws3.cell(r, 2, "").border = bdr
    ref_cost_row = cost_end_row + 1
    ws3.cell(r, 3, f"=C{ref_cost_row}").number_format = "$#,##0.00"; ws3.cell(r, 3).border = bdr; ws3.cell(r, 3).alignment = right

    r += 1
    ws3.row_dimensions[r].height = 22
    net_cell_row = r
    ws3.cell(r, 1, "Net Profit / (Loss) on Delivery").font = Font(bold=True, size=10)
    ws3.cell(r, 1).border = Border(top=thick_side, bottom=thick_side, left=thin, right=thin)
    ws3.cell(r, 1).alignment = left
    ws3.cell(r, 2, "").border = Border(top=thick_side, bottom=thick_side, left=thin, right=thin)
    net_c = ws3.cell(r, 3)
    net_c.value = f"={delivery_charge}-C{ref_cost_row}"
    net_c.number_format = "$#,##0.00"
    net_c.font = Font(bold=True, size=10)
    net_c.border = Border(top=thick_side, bottom=thick_side, left=thin, right=thin)
    net_c.alignment = right
    # Note: colour is static since formula is dynamic — add a helper note
    r += 1
    ws3.cell(r, 1, "ℹ  Positive = profit on delivery  |  Negative = loss on delivery").font = Font(italic=True, size=8, color="888888")
    ws3.cell(r, 1).alignment = left

    xlsx_path = Path(output_path).with_name(Path(output_path).stem + "_INTERNAL.xlsx")
    wb.save(xlsx_path)
    return str(xlsx_path)


def draw_invoice(inv, output_path):
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=0.28 * inch,
        rightMargin=0.28 * inch,
        topMargin=1.55 * inch,
        bottomMargin=1.00 * inch,
    )
    doc._gg_invoice = inv

    elements = []
    temp_images = []

    label_style = ParagraphStyle(
        "label_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=DARK,
        leading=10,
    )
    value_style = ParagraphStyle(
        "value_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=BLACK,
        leading=10,
    )
    cell_style = ParagraphStyle(
        "cell_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=BLACK,
        leading=10,
    )
    note_style = ParagraphStyle(
        "note_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.3,
        textColor=DARK,
        leading=10,
    )
    notes_bold_style = ParagraphStyle(
        "notes_bold_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        textColor=DARK,
        leading=11,
    )
    invoice_title_style = ParagraphStyle(
        "invoice_title_style",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        textColor=BLACK,
        alignment=1,
        spaceAfter=6,
    )

    # Logo is drawn by _NumberedCanvas on every page (see _draw_footer_and_header)

    client_rows = [
        [Paragraph("Tel. | Mob.:", label_style), Paragraph(inv.get("client_phone", ""), value_style)],
        [Paragraph("Name:", label_style), Paragraph(inv.get("client_name", ""), value_style)],
    ]

    if inv.get("client_email"):
        client_rows.append([Paragraph("Email:", label_style), Paragraph(inv.get("client_email", ""), value_style)])

    for idx, line in enumerate(inv.get("client_address", []) or []):
        client_rows.append([
            Paragraph("Del. Address:" if idx == 0 else "", label_style),
            Paragraph(line, value_style),
        ])

    client_tbl = Table(client_rows, colWidths=[1.05 * inch, CONTENT_WIDTH - 1.05 * inch])
    client_tbl.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(client_tbl)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("Invoice", invoice_title_style))

    # EXACT same total width as item table
    meta_widths = [1.45 * inch, 1.40 * inch, 1.35 * inch, CONTENT_WIDTH - 1.45 * inch - 1.40 * inch - 1.35 * inch]
    meta_data = [
        ["Invoice Date:", "Invoice", "Client No:", "Your Reference:"],
        [inv.get("date", ""), inv.get("number", ""), str(inv.get("client_no", "")), inv.get("reference", "")],
    ]
    meta_tbl = Table(meta_data, colWidths=meta_widths)
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLACK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 12))

    headers = ["ITEM NO.", "DESCRIPTION", "EST. DELIVERY", "TYPE", "QTY", "UNIT PRICE", "DISC.", "TOTAL", "PHOTO"]
    rows = [headers]

    for item in inv.get("items", []):
        qty = item.get("qty", 0)
        unit_price = item.get("unit_price", 0)
        disc = item.get("discount", 0)
        line_total = qty * unit_price * (1 - disc)

        photo_cell = ""
        img_path = _decode_image(item.get("image", ""))

        if img_path:
            temp_images.append(img_path)
            try:
                img = Image(img_path, width=0.96 * inch, height=0.70 * inch)
                img.hAlign = "CENTER"
                photo_cell = img
            except Exception:
                photo_cell = ""

        rows.append([
            Paragraph(str(item.get("no", "")), cell_style),
            Paragraph(item.get("description", ""), cell_style),
            Paragraph(item.get("delivery", ""), cell_style),
            Paragraph(item.get("unit", ""), cell_style),
            Paragraph(str(qty), cell_style),
            Paragraph(usd(unit_price), cell_style),
            Paragraph(f"{disc * 100:.0f}%" if disc else "", cell_style),
            Paragraph(usd(line_total), cell_style),
            photo_cell,
        ])

    # Calculate row heights: fixed header, bounded data rows
    HDR_H   = 0.45 * inch
    DATA_H  = 0.85 * inch   # max row height — prevents runaway cell growth
    row_heights = [HDR_H] + [DATA_H] * (len(rows) - 1)
    item_tbl = Table(rows, colWidths=ITEM_COL_WIDTHS,
                     rowHeights=row_heights, repeatRows=1)
    item_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLACK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-2, -1), 4),
        ("RIGHTPADDING", (0, 0), (-2, -1), 4),
        ("LEFTPADDING", (-1, 1), (-1, -1), 2),
        ("RIGHTPADDING", (-1, 1), (-1, -1), 2),
        ("ALIGN", (3, 1), (8, -1), "CENTER"),
    ]))
    elements.append(item_tbl)
    elements.append(Spacer(1, 10))

    subtotal = sum(i.get("qty", 0) * i.get("unit_price", 0) * (1 - i.get("discount", 0)) for i in inv.get("items", []))
    delivery_charge = inv.get("delivery_charge", 0) or 0
    tax_amt = (subtotal + delivery_charge) * (inv.get("tax_rate", 0) or 0)
    total = subtotal + delivery_charge + tax_amt

    # Right aligned delivery label
    if inv.get("delivery_type"):
        delivery_tbl = Table([[Paragraph(f"<b>{inv.get('delivery_type')}</b>", cell_style)]], colWidths=[CONTENT_WIDTH])
        delivery_tbl.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(delivery_tbl)
        elements.append(Spacer(1, 4))

    totals_tbl = Table([
        ["SubTotal", usd(subtotal)],
        ["Delivery Charge", usd(delivery_charge)],
        ["Sales Tax", usd(tax_amt)],
        ["Total", usd(total)],
    ], colWidths=[1.7 * inch, 1.0 * inch])
    totals_tbl.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 0.8, BLACK),
        ("LINEBELOW", (0, 2), (-1, 2), 0.8, BORDER),
        ("LINEBELOW", (0, 3), (-1, 3), 0.8, BLACK),
        ("FONTNAME", (0, 0), (-1, -2), "Helvetica"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    totals_wrap = Table([["", totals_tbl]], colWidths=[CONTENT_WIDTH - 2.70 * inch, 2.70 * inch])
    totals_wrap.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(totals_wrap)
    elements.append(Spacer(1, 12))

    # ── Payment plan (installments) ───────────────────────────────────────────
    installments = inv.get("installments", [])
    if installments and inv.get("payment_terms") == "installments":
        split_type = inv.get("installment_split_type", "amount")
        # Calculate amounts, force last installment = total - sum(others)
        # so the payment plan total ALWAYS exactly matches the invoice total
        amts = []
        for inst in installments:
            val = float(inst.get("val", 0) or 0)
            if split_type == "pct":
                amts.append(round(total * (val / 100.0), 2))
            else:
                amts.append(round(val, 2))
        if amts:
            amts[-1] = round(total - sum(amts[:-1]), 2)  # exact match, no rounding error

        # Build rows: header + each installment + exact total
        plan_rows = [["Payment Terms", "", ""]]
        for i, (inst, amt) in enumerate(zip(installments, amts)):
            label = f"{i+1}{'st' if i==0 else 'nd' if i==1 else 'rd' if i==2 else 'th'} Installment"
            plan_rows.append([label, inst.get("date", ""), usd(amt)])
        plan_rows.append(["", "", usd(total)])  # always exactly the invoice total

        plan_tbl = Table(plan_rows, colWidths=[1.4*inch, 1.1*inch, 0.85*inch])
        plan_style = [
            ("FONTNAME",   (0,0), (-1,0),   "Helvetica-Bold"),
            ("FONTNAME",   (0,1), (-1,-2),  "Helvetica"),
            ("FONTNAME",   (0,-1),(-1,-1),  "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1),  8),
            ("ALIGN",      (1,0), (1,-1),   "CENTER"),
            ("ALIGN",      (2,0), (2,-1),   "RIGHT"),
            ("TOPPADDING", (0,0), (-1,-1),  3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ("LINEABOVE",  (0,0), (-1,0),   0.6, colors.HexColor("#231f1e")),
            ("LINEBELOW",  (0,-2),(-1,-2),  0.6, colors.HexColor("#d8d5d2")),
            ("LINEBELOW",  (0,-1),(-1,-1),  0.6, colors.HexColor("#231f1e")),
            ("SPAN",       (0,0), (1,0)),
        ]
        plan_tbl.setStyle(TableStyle(plan_style))

        plan_wrap = Table([["", plan_tbl]],
                          colWidths=[CONTENT_WIDTH - 3.35*inch, 3.35*inch])
        plan_wrap.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 0),
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ]))
        elements.append(plan_wrap)
        elements.append(Spacer(1, 10))

    # ── Notes (from invoice form) ──────────────────────────────────
    if inv.get("notes", "").strip():
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(inv["notes"].strip(), notes_bold_style))
        elements.append(Spacer(1, 4))

    # Fixed payment line always shown, then bold payment method on same line
    payment_method = inv.get("payment_terms", "advance")
    bold_text = "Paid in advance." if payment_method != "installments" else "Paid in installments."
    pay_line = (
        f"Payment is via check, bank transfer, or credit card. "
        f"Please note that credit card payments incur a 3% processing fee. "
        f"<b>{bold_text}</b>"
    )
    elements.append(Paragraph(pay_line, note_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "Bank account details #: 930283558  Routing number: 061092387.  Zelle email: rana_salah@goldenglam.nl",
        note_style
    ))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        "All quote(s), (provisional) order(s) (confirmations), sales and deliveries are subject to the Golden Glam Terms of orders and payments, the Golden Glam Reseller Terms and the CBM General Sales Terms and Conditions. US law applies.",
        note_style
    ))

    _NumberedCanvas._gg_inv = inv  # make invoice accessible inside canvas callbacks
    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer,
              canvasmaker=_NumberedCanvas)

    xlsx_path = _write_internal_excel(inv, output_path)

    for img in temp_images:
        try:
            os.remove(img)
        except Exception:
            pass

    return xlsx_path
